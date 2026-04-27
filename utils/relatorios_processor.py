"""
utils/relatorios_processor.py — Gerador de relatórios Excel analíticos de
pagamentos autorizados para o Robô Financeiro BOAH/SOLAR.
"""
from __future__ import annotations

import os
from datetime import datetime
from collections import defaultdict


def gerar_excel_pagamentos_autorizados(caminho_saida: str) -> tuple[bool, str]:
    """
    Lê a tabela `pagamentos_autorizados` do banco e gera um arquivo Excel com
    as seguintes abas analíticas:

        1. Resumo Mensal   — soma por mês (YYYY-MM)
        2. Resumo Semanal  — soma por semana do ano (YYYY-Wxx)
        3. Resumo Diário   — soma por dia (DD/MM/YYYY)
        4. Por Categoria   — soma por categoria contábil
        5. Dados Brutos    — histórico completo para tabela dinâmica

    Retorna (True, "ok") em caso de sucesso ou (False, mensagem_de_erro).
    """
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False, (
            "openpyxl não encontrado. Execute: pip install openpyxl"
        )

    try:
        from database import listar_pagamentos_autorizados
    except ImportError as e:
        return False, f"Não foi possível importar o banco de dados: {e}"

    # ── 1. Carrega dados ────────────────────────────────────────────────────
    registros = listar_pagamentos_autorizados()

    if not registros:
        return False, (
            "Nenhum pagamento encontrado no banco. "
            "Gere ao menos um PDF de Autorização primeiro."
        )

    # ── 2. Agrega os dados ──────────────────────────────────────────────────
    por_mes      = defaultdict(float)
    por_semana   = defaultdict(float)
    por_dia      = defaultdict(float)
    por_categoria = defaultdict(float)

    for r in registros:
        valor = r.get("valor", 0.0) or 0.0
        dt = r.get("data_obj")

        if dt:
            por_mes[dt.strftime("%Y-%m")] += valor
            semana = f"{dt.year}-W{dt.isocalendar()[1]:02d}"
            por_semana[semana] += valor
            por_dia[dt.strftime("%d/%m/%Y")] += valor

        cat = r.get("categoria") or "A CLASSIFICAR"
        por_categoria[cat] += valor

    # ── 3. Estilos comuns ───────────────────────────────────────────────────
    COR_HDR_BG  = "0F172A"   # azul-escuro (cabeçalho)
    COR_HDR_FG  = "E2E8F0"   # cinza-claro (texto cabeçalho)
    COR_ALT     = "1E293B"   # alternância de linhas (escuro)
    COR_ALT2    = "0A0F1E"   # linha par
    COR_FG      = "E2E8F0"   # texto padrão
    COR_TOTAL   = "064E3B"   # verde-escuro (linha de total)
    COR_TOTAL_FG= "4ADE80"   # verde-claro (texto do total)
    COR_ACCENT  = "38BDF8"   # azul accent

    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def _font(hex_fg: str = COR_FG, bold: bool = False,
              size: int = 10) -> Font:
        return Font(color=hex_fg, bold=bold, size=size, name="Segoe UI")

    def _border() -> Border:
        thin = Side(style="thin", color="334155")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _align(horizontal: str = "left") -> Alignment:
        return Alignment(horizontal=horizontal, vertical="center", wrap_text=False)

    def _fmt_currency(ws, cell):
        cell.number_format = '#,##0.00'

    def _aplicar_header(ws, colunas: list[str], larguras: list[int]):
        ws.append(colunas)
        for col_idx, (col_name, width) in enumerate(zip(colunas, larguras), start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill      = _fill(COR_HDR_BG)
            cell.font      = _font(COR_HDR_FG, bold=True, size=11)
            cell.alignment = _align("center")
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 22

    def _preencher_linhas(ws, linhas: list[tuple], valor_cols: set[int]):
        """
        Insere linhas de dados com formatação alternada.
        valor_cols: índices (1-based) das colunas de valor monetário.
        """
        for row_idx, linha in enumerate(linhas, start=2):
            bg = COR_ALT if row_idx % 2 == 0 else COR_ALT2
            for col_idx, val in enumerate(linha, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill      = _fill(bg)
                cell.font      = _font()
                cell.alignment = _align(
                    "right" if col_idx in valor_cols else "left"
                )
                cell.border    = _border()
                if col_idx in valor_cols and isinstance(val, (int, float)):
                    cell.number_format = '"R$" #,##0.00'

    def _linha_total(ws, row_num: int, label: str, total: float,
                     label_col: int, valor_col: int, num_cols: int):
        """Adiciona linha de TOTAL ao final da aba."""
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill   = _fill(COR_TOTAL)
            cell.border = _border()
            cell.font   = _font(COR_TOTAL_FG, bold=True, size=11)
            if c == label_col:
                cell.value     = label
                cell.alignment = _align("left")
            elif c == valor_col:
                cell.value        = total
                cell.number_format = '"R$" #,##0.00'
                cell.alignment    = _align("right")
            else:
                cell.value = ""
        
        # Nota de rodapé da marca
        ws.cell(row=row_num+2, column=1, value="Gerado por Com System - Agência de Atendimento Digital")
        ws.cell(row=row_num+2, column=1).font = _font(hex_fg="525252", size=8)

    # ── 4. Cria o workbook ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove a aba padrão vazia

    # ── ABA: Resumo Mensal ──────────────────────────────────────────────────
    ws_mes = wb.create_sheet("📅 Resumo Mensal")
    _aplicar_header(ws_mes, ["Mês", "Total (R$)", "Nº Pagamentos"], [18, 20, 18])
    contagem_mes = defaultdict(int)
    for r in registros:
        dt = r.get("data_obj")
        if dt:
            contagem_mes[dt.strftime("%Y-%m")] += 1
    linhas_mes = sorted(por_mes.items())
    _preencher_linhas(
        ws_mes,
        [(mes, valor, contagem_mes[mes]) for mes, valor in linhas_mes],
        {2}
    )
    total_geral = sum(por_mes.values())
    _linha_total(ws_mes, len(linhas_mes) + 2, "TOTAL GERAL",
                 total_geral, 1, 2, 3)

    # ── ABA: Resumo Semanal ─────────────────────────────────────────────────
    ws_sem = wb.create_sheet("📆 Resumo Semanal")
    _aplicar_header(ws_sem, ["Semana", "Total (R$)", "Nº Pagamentos"], [18, 20, 18])
    contagem_sem = defaultdict(int)
    for r in registros:
        dt = r.get("data_obj")
        if dt:
            semana = f"{dt.year}-W{dt.isocalendar()[1]:02d}"
            contagem_sem[semana] += 1
    linhas_sem = sorted(por_semana.items())
    _preencher_linhas(
        ws_sem,
        [(sem, valor, contagem_sem[sem]) for sem, valor in linhas_sem],
        {2}
    )
    _linha_total(ws_sem, len(linhas_sem) + 2, "TOTAL GERAL",
                 total_geral, 1, 2, 3)

    # ── ABA: Resumo Diário ──────────────────────────────────────────────────
    ws_dia = wb.create_sheet("📊 Resumo Diário")
    _aplicar_header(ws_dia, ["Data", "Total (R$)", "Nº Pagamentos"], [16, 20, 18])
    contagem_dia = defaultdict(int)
    for r in registros:
        dt = r.get("data_obj")
        if dt:
            contagem_dia[dt.strftime("%d/%m/%Y")] += 1

    # Ordena por data real (não string)
    datas_ordenadas = sorted(por_dia.items(),
                             key=lambda x: datetime.strptime(x[0], "%d/%m/%Y"))
    _preencher_linhas(
        ws_dia,
        [(dia, valor, contagem_dia[dia]) for dia, valor in datas_ordenadas],
        {2}
    )
    _linha_total(ws_dia, len(datas_ordenadas) + 2, "TOTAL GERAL",
                 total_geral, 1, 2, 3)

    # ── ABA: Por Categoria ──────────────────────────────────────────────────
    ws_cat = wb.create_sheet("🏷️ Por Categoria")
    _aplicar_header(ws_cat, ["Categoria Contábil", "Total (R$)",
                              "Nº Pagamentos", "% do Total"], [45, 20, 18, 14])
    contagem_cat = defaultdict(int)
    for r in registros:
        cat = r.get("categoria") or "A CLASSIFICAR"
        contagem_cat[cat] += 1

    linhas_cat = sorted(por_categoria.items(),
                        key=lambda x: x[1], reverse=True)
    rows_cat = []
    for cat, valor in linhas_cat:
        pct = (valor / total_geral * 100) if total_geral else 0.0
        rows_cat.append((cat, valor, contagem_cat[cat], round(pct, 2)))
    _preencher_linhas(ws_cat, rows_cat, {2, 4})

    # Formata coluna % com símbolo
    for row_idx in range(2, len(rows_cat) + 2):
        cell_pct = ws_cat.cell(row=row_idx, column=4)
        cell_pct.number_format = '0.00"%"'

    _linha_total(ws_cat, len(linhas_cat) + 2, "TOTAL GERAL",
                 total_geral, 1, 2, 4)

    # ── ABA: Dados Brutos ───────────────────────────────────────────────────
    ws_raw = wb.create_sheet("📋 Dados Brutos")
    colunas_raw = ["ID", "Nome/Beneficiário", "CNPJ", "Tipo", "Data",
                   "Valor (R$)", "Status", "Responsável", "Categoria",
                   "Empresa", "Observação", "Data Aprovação"]
    larguras_raw = [6, 38, 18, 22, 12, 16, 12, 20, 38, 10, 30, 20]
    _aplicar_header(ws_raw, colunas_raw, larguras_raw)

    linhas_raw = []
    for r in registros:
        linhas_raw.append((
            r.get("id", ""),
            r.get("nome", ""),
            r.get("cnpj", ""),
            r.get("tipo", ""),
            r.get("data", ""),
            r.get("valor", 0.0),
            r.get("status", ""),
            r.get("responsavel", ""),
            r.get("categoria", ""),
            r.get("empresa", ""),
            r.get("observacao", ""),
            r.get("data_aprovacao", ""),
        ))
    _preencher_linhas(ws_raw, linhas_raw, {6})

    # Congela a primeira linha em todas as abas
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        # Cor de fundo da aba
        ws.sheet_properties.tabColor = COR_ACCENT.lower()

    # ── 5. Salva ────────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(caminho_saida) or ".", exist_ok=True)
    wb.save(caminho_saida)

    return True, f"{len(registros)} registros exportados com sucesso."
